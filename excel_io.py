"""
excel_io.py — Excel 讀檔層（純 I/O，無任何清洗邏輯）

v1.1（F-3）新增：從 boq_cleaner.py 抽出 read_grid() / list_sheets()。

抽出原因：quote_cleaner.py 原本為了重用這兩個函式而 import 整份
boq_cleaner.py（474 行，含全部標單清洗邏輯：classify_itemno、clean_boq、
_choose_qty 等）。這造成兩個問題：
  1. 報價單批次包必須夾帶一份完整的標單核心，體積與認知負擔都不必要。
  2. 未來若只更新其中一個包的 boq_cleaner.py，會出現兩個版本的
     boq_cleaner.py 同時流通，使用者無從分辨（版本漂移風險）。

抽出後：boq_cleaner.py 與 quote_cleaner.py 都改成從這個約 30 行的小檔案
import，報價單批次包只需要帶 excel_io.py，不用再帶整份 boq_cleaner.py。

本檔不包含、也不應該包含任何清洗規則——只負責把 .xls / .xlsx 讀成
Python 原生的二維 list，交給呼叫端各自的清洗核心處理。
"""


def read_grid(path, sheet_name):
    """讀取指定工作表為二維 list（保留原始儲存格值，不做任何清洗）。"""
    if str(path).lower().endswith('.xls'):
        import xlrd
        wb = xlrd.open_workbook(path)
        ws = wb.sheet_by_name(sheet_name)
        return [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]
    else:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet_name]
        return [list(row) for row in ws.iter_rows(values_only=True)]


def list_sheets(path):
    """回傳活頁簿內所有工作表名稱。"""
    if str(path).lower().endswith('.xls'):
        import xlrd
        return xlrd.open_workbook(path).sheet_names()
    else:
        from openpyxl import load_workbook
        return load_workbook(path, read_only=True).sheetnames
